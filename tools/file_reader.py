"""ファイルコンテンツ抽出ツール - 複数のファイル形式に対応"""
import json
from pathlib import Path
from typing import Optional


def read_text_file(file_path: str, encoding: str = "utf-8") -> str:
    """プレーンテキストファイルを読み取り（TXT, MD, XML, HTML, CSV, LOGなど）"""
    path = Path(file_path)
    try:
        return path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        # 他のエンコーディングを試行
        for enc in ["utf-8-sig", "gbk", "gb2312", "shift_jis", "latin-1"]:
            try:
                return path.read_text(encoding=enc)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"ファイルをデコードできません: {file_path}")


def read_json_file(file_path: str) -> str:
    """JSONファイルを読み取り、フォーマットして出力"""
    path = Path(file_path)
    content = path.read_text(encoding="utf-8")
    data = json.loads(content)
    return json.dumps(data, indent=2, ensure_ascii=False)


def read_pdf_file(file_path: str) -> str:
    """PyMuPDFを使用してPDFテキストを抽出"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError("PyMuPDFのインストールが必要です: pip install PyMuPDF")

    doc = fitz.open(file_path)
    pages = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()
        if text.strip():
            pages.append(f"--- Page {page_num + 1} ---\n{text}")
    doc.close()
    return "\n\n".join(pages)


def read_docx_file(file_path: str) -> str:
    """python-docxを使用してWordドキュメントのテキストを抽出"""
    try:
        from docx import Document
    except ImportError:
        raise ImportError("python-docxのインストールが必要です: pip install python-docx")

    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def read_xlsx_file(file_path: str) -> str:
    """openpyxlを使用してExcelファイルを読み取り"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxlのインストールが必要です: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows.append("\t".join(cells))
        if rows:
            sheets.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    wb.close()
    return "\n\n".join(sheets)


# 対応するファイル拡張子と読み取り関数のマッピング
TEXT_EXTENSIONS = {".txt", ".md", ".xml", ".html", ".htm", ".csv", ".log",
                   ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
                   ".py", ".js", ".ts", ".java", ".c", ".cpp", ".h",
                   ".go", ".rs", ".rb", ".php", ".sql", ".sh", ".bat"}


def read_file(file_path: str) -> str:
    """
    ファイル拡張子に基づいて自動的に読み取り方式を選択

    Args:
        file_path: ファイルパス

    Returns:
        ファイルコンテンツのテキスト
    """
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".json":
        return read_json_file(file_path)
    elif ext == ".pdf":
        return read_pdf_file(file_path)
    elif ext == ".docx":
        return read_docx_file(file_path)
    elif ext in (".xlsx", ".xls"):
        return read_xlsx_file(file_path)
    elif ext in TEXT_EXTENSIONS or ext == "":
        return read_text_file(file_path)
    else:
        # テキストファイルとして読み取りを試行
        try:
            return read_text_file(file_path)
        except Exception:
            raise ValueError(f"対応していないファイル形式: {ext}")


def get_supported_extensions() -> list:
    """対応するファイル拡張子リストを返却"""
    return sorted(list(TEXT_EXTENSIONS | {".json", ".pdf", ".docx", ".xlsx", ".xls"}))
